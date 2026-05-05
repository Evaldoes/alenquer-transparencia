"""
Exportação de dados em CSV e Excel.
"""
import csv
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fmt_brl(v):
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(v)


def _nivel(score):
    if score >= 70: return "ALTO"
    if score >= 40: return "MÉDIO"
    return "BAIXO"


# ── CSV ───────────────────────────────────────────────────────────────────────

def exportar_csv(resultados_monitor):
    """Retorna bytes de um CSV com todos os contratos analisados."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_ALL)

    writer.writerow([
        "Score", "Nível de Risco", "Empresa", "CNPJ",
        "Objeto", "Valor", "Modalidade",
        "Empresa Aberta Em", "Capital Social", "Porte",
        "Sanção CEIS", "Sanção CNEP",
        "Flags", "Analisado Em",
    ])

    for r in sorted(resultados_monitor, key=lambda x: x.get("score", 0), reverse=True):
        c  = r.get("contrato", {})
        cd = r.get("cnpj_dados") or {}
        s  = r.get("sancoes") or {}
        writer.writerow([
            r.get("score", 0),
            _nivel(r.get("score", 0)),
            c.get("nomeContratado", ""),
            c.get("cnpjContratado", ""),
            str(c.get("objetoCompra") or "")[:200],
            _fmt_brl(c.get("valorInicialCompra") or c.get("valor") or 0),
            c.get("modalidadeCompra", ""),
            cd.get("inicio", ""),
            _fmt_brl(cd.get("capital") or 0),
            cd.get("porte", ""),
            len(s.get("ceis", [])),
            len(s.get("cnep", [])),
            " | ".join(r.get("flags", [])[:5]),
            r.get("analisado_em", "")[:10],
        ])

    return buf.getvalue().encode("utf-8-sig")  # BOM para Excel abrir corretamente


# ── Excel ─────────────────────────────────────────────────────────────────────

_COR_ALTO    = "FC8181"
_COR_MEDIO   = "F6AD55"
_COR_BAIXO   = "68D391"
_COR_HEADER  = "1A6B3C"
_COR_BRANCO  = "FFFFFF"
_COR_CINZA   = "EDF2F7"


def _header_style(cell):
    cell.font      = Font(bold=True, color=_COR_BRANCO, size=10)
    cell.fill      = PatternFill("solid", fgColor=_COR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _cor_score(score):
    if score >= 70: return _COR_ALTO
    if score >= 40: return _COR_MEDIO
    return _COR_BAIXO


def exportar_excel(resultados_monitor, pares_copypaste=None, dados_pncp=None):
    """
    Gera um arquivo Excel com múltiplas abas:
      - Contratos analisados (com score de risco)
      - Copy-paste detectado
      - PNCP (contratações recentes)
      - Resumo executivo
    Retorna bytes prontos para download.
    """
    wb = openpyxl.Workbook()

    # ── Aba 1: Contratos analisados ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Contratos Analisados"

    cabecalhos = [
        "Score", "Risco", "Empresa", "CNPJ", "Objeto",
        "Valor (R$)", "Modalidade", "Abertura Empresa",
        "Capital", "Porte", "CEIS", "CNEP", "Flags",
    ]
    ws.append(cabecalhos)
    for cell in ws[1]:
        _header_style(cell)
    ws.row_dimensions[1].height = 28

    for r in sorted(resultados_monitor, key=lambda x: x.get("score", 0), reverse=True):
        c  = r.get("contrato", {})
        cd = r.get("cnpj_dados") or {}
        s  = r.get("sancoes") or {}
        score = r.get("score", 0)
        linha = [
            score,
            _nivel(score),
            c.get("nomeContratado", ""),
            c.get("cnpjContratado", ""),
            str(c.get("objetoCompra") or "")[:200],
            float(c.get("valorInicialCompra") or c.get("valor") or 0),
            c.get("modalidadeCompra", ""),
            cd.get("inicio", ""),
            float(cd.get("capital") or 0),
            cd.get("porte", ""),
            len(s.get("ceis", [])),
            len(s.get("cnep", [])),
            " | ".join(r.get("flags", [])[:4]),
        ]
        ws.append(linha)
        row_n = ws.max_row
        cor   = _cor_score(score)
        for col in range(1, 3):
            ws.cell(row_n, col).fill = PatternFill("solid", fgColor=cor)
            ws.cell(row_n, col).font = Font(bold=True, size=9)
        for col in range(3, len(linha) + 1):
            ws.cell(row_n, col).font      = Font(size=9)
            ws.cell(row_n, col).alignment = Alignment(wrap_text=True, vertical="top")
            if row_n % 2 == 0:
                ws.cell(row_n, col).fill = PatternFill("solid", fgColor=_COR_CINZA)
        # Formatar valor como moeda
        ws.cell(row_n, 6).number_format = 'R$ #,##0.00'
        ws.cell(row_n, 9).number_format = 'R$ #,##0.00'

    # Larguras das colunas
    larguras = [7, 9, 35, 18, 60, 16, 22, 14, 14, 10, 6, 6, 80]
    for i, l in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = l

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Aba 2: Copy-paste ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Copy-Paste Detectado")
    ws2.append(["Similaridade %", "Empresa A", "Empresa B", "Objeto A", "Objeto B"])
    for cell in ws2[1]:
        _header_style(cell)

    for par in (pares_copypaste or []):
        ws2.append([
            par.get("similaridade", 0),
            par.get("contrato_a", {}).get("nomeContratado", ""),
            par.get("contrato_b", {}).get("nomeContratado", ""),
            par.get("objeto_a", ""),
            par.get("objeto_b", ""),
        ])
        row_n = ws2.max_row
        sim   = par.get("similaridade", 0)
        cor   = _COR_ALTO if sim >= 90 else _COR_MEDIO
        ws2.cell(row_n, 1).fill = PatternFill("solid", fgColor=cor)
        ws2.cell(row_n, 1).font = Font(bold=True, size=9)
        for col in range(2, 6):
            ws2.cell(row_n, col).font = Font(size=9)
            ws2.cell(row_n, col).alignment = Alignment(wrap_text=True)

    for i, l in enumerate([14, 35, 35, 60, 60], 1):
        ws2.column_dimensions[get_column_letter(i)].width = l

    # ── Aba 3: PNCP ─────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("PNCP — Contratos Recentes")
    ws3.append(["Empresa", "CNPJ", "Objeto", "Valor (R$)", "Modalidade", "Data"])
    for cell in ws3[1]:
        _header_style(cell)

    for c in (dados_pncp or []):
        ws3.append([
            c.get("nomeContratado", ""),
            c.get("cnpjContratado", ""),
            str(c.get("objetoCompra") or "")[:200],
            float(c.get("valorInicialCompra") or 0),
            c.get("modalidadeCompra", ""),
            c.get("dataAssinatura", "")[:10],
        ])
        ws3.cell(ws3.max_row, 4).number_format = 'R$ #,##0.00'

    for i, l in enumerate([35, 18, 80, 16, 22, 12], 1):
        ws3.column_dimensions[get_column_letter(i)].width = l

    # ── Aba 4: Resumo ────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Resumo Executivo")
    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 20

    alto  = sum(1 for r in resultados_monitor if r.get("score", 0) >= 70)
    medio = sum(1 for r in resultados_monitor if 40 <= r.get("score", 0) < 70)
    total_valor = sum(
        float(r.get("contrato", {}).get("valorInicialCompra") or 0)
        for r in resultados_monitor
    )

    linhas_resumo = [
        ("Gerado em",            datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Município",            "Alenquer / PA"),
        ("",                     ""),
        ("Contratos analisados", len(resultados_monitor)),
        ("Alto risco (≥70)",     alto),
        ("Médio risco (40–69)",  medio),
        ("Baixo risco (<40)",    len(resultados_monitor) - alto - medio),
        ("Valor total",          _fmt_brl(total_valor)),
        ("",                     ""),
        ("Copy-paste detectado", len(pares_copypaste or [])),
        ("Contratos PNCP",       len(dados_pncp or [])),
    ]
    for label, valor in linhas_resumo:
        ws4.append([label, valor])
        if label:
            ws4.cell(ws4.max_row, 1).font = Font(bold=True, size=10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
